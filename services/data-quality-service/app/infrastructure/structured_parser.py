"""Bộ phân tích cú pháp dữ liệu có cấu trúc (UC-029, bước 3-4).

`parse_raw_bytes()`: đọc bytes thô theo định dạng (CSV/EXCEL/JSON/XML) ->
danh sách dict {tên_cột_nguồn: giá_trị_chuỗi_hoặc_thô}. Đây là bước 3
"Phân tích Excel/CSV/JSON/XML theo lược đồ".

`map_and_cast_row()`: ánh xạ tên trường (theo `field_mapping` tường minh
hoặc tự động khớp tên) + ép kiểu theo `schema_fields` (bước 4).

Thuần Python, không phụ thuộc DB/HTTP — dễ unit test độc lập.
"""
from __future__ import annotations

import csv
import io
import json
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime
from typing import Any, Dict, List, Tuple


class ParseError(Exception):
    """Lỗi không đọc/phân tích được toàn bộ nội dung thô (không phải lỗi
    ép kiểu từng dòng — lỗi từng dòng được thu thập riêng, không raise)."""


def parse_raw_bytes(content: bytes, source_format: str) -> List[Dict[str, Any]]:
    """Bước 3: phân tích nội dung thô -> danh sách dict theo từng dòng/bản ghi."""
    if source_format == "CSV":
        return _parse_csv(content)
    if source_format == "JSON":
        return _parse_json(content)
    if source_format == "XML":
        return _parse_xml(content)
    if source_format == "EXCEL":
        return _parse_excel(content)
    raise ParseError(f"Định dạng không được hỗ trợ: {source_format}")


def _parse_csv(content: bytes) -> List[Dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ParseError(f"Không giải mã được nội dung CSV (utf-8): {exc}") from exc
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_json(content: bytes) -> List[Dict[str, Any]]:
    try:
        data = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ParseError(f"Không phân tích được JSON: {exc}") from exc
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        # Cho phép payload dạng {"records": [...]}/{"data": [...]} — tìm
        # danh sách bản ghi đầu tiên trong dict.
        rows = next((v for v in data.values() if isinstance(v, list)), None)
        if rows is None:
            raise ParseError("JSON dạng object phải chứa 1 danh sách bản ghi")
    else:
        raise ParseError("JSON phải là danh sách bản ghi hoặc object chứa danh sách bản ghi")
    if not all(isinstance(r, dict) for r in rows):
        raise ParseError("Mỗi bản ghi JSON phải là 1 object")
    return rows


def _parse_xml(content: bytes) -> List[Dict[str, Any]]:
    try:
        root = ET.fromstring(content)
    except ET.ParseError as exc:
        raise ParseError(f"Không phân tích được XML: {exc}") from exc
    rows: List[Dict[str, Any]] = []
    # Quy ước: phần tử gốc chứa danh sách phần tử con lặp lại, mỗi phần tử
    # con đại diện 1 bản ghi, các phần tử cháu là các trường (tag: text).
    for record_el in root:
        row: Dict[str, Any] = {}
        for field_el in record_el:
            row[field_el.tag] = field_el.text
        if row:
            rows.append(row)
    return rows


def _parse_excel(content: bytes) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError("Thiếu thư viện 'openpyxl' để đọc tệp Excel") from exc
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - lỗi định dạng tệp Excel bất kỳ
        raise ParseError(f"Không đọc được tệp Excel: {exc}") from exc
    sheet = wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    rows: List[Dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        rows.append({header[i]: values[i] for i in range(len(header)) if i < len(values)})
    return rows


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", "_", str(name).strip().lower())


def build_auto_field_mapping(
    raw_columns: List[str], schema_fields: List[Dict[str, Any]]
) -> Dict[str, str]:
    """Tự động ánh xạ tên cột nguồn -> tên trường đích khi không có
    `field_mapping` tường minh: khớp theo tên đã chuẩn hoá (bỏ khoảng
    trắng thừa, không phân biệt hoa/thường)."""
    target_by_normalized = {_normalize_name(f["name"]): f["name"] for f in schema_fields}
    mapping: Dict[str, str] = {}
    for col in raw_columns:
        target = target_by_normalized.get(_normalize_name(col))
        if target:
            mapping[col] = target
    return mapping


def cast_value(raw_value: Any, data_type: str) -> Any:
    """Ép kiểu 1 giá trị thô theo `data_type` của lược đồ đích. Raise
    `ValueError` nếu không ép được — caller (application layer) bắt lỗi
    này để ghi vào `parsing_row_errors` (không làm hỏng cả dòng/cả job)."""
    if raw_value is None:
        return None
    if isinstance(raw_value, str) and raw_value.strip() == "":
        return None

    if data_type == "STRING":
        return str(raw_value)

    if data_type in ("INTEGER", "BIGINT"):
        if isinstance(raw_value, bool):
            raise ValueError(f"Không thể ép kiểu bool '{raw_value}' thành {data_type}")
        return int(str(raw_value).strip())

    if data_type == "DECIMAL":
        return float(str(raw_value).strip())

    if data_type == "BOOLEAN":
        if isinstance(raw_value, bool):
            return raw_value
        text = str(raw_value).strip().lower()
        if text in ("true", "1", "yes", "y", "có", "đúng"):
            return True
        if text in ("false", "0", "no", "n", "không", "sai"):
            return False
        raise ValueError(f"Không thể ép kiểu '{raw_value}' thành BOOLEAN")

    if data_type == "DATE":
        if isinstance(raw_value, (date, datetime)):
            return raw_value.isoformat()[:10]
        text = str(raw_value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        raise ValueError(f"Không thể ép kiểu '{raw_value}' thành DATE (kỳ vọng YYYY-MM-DD)")

    if data_type == "DATETIME":
        if isinstance(raw_value, datetime):
            return raw_value.isoformat()
        text = str(raw_value).strip()
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).isoformat()
            except ValueError:
                continue
        raise ValueError(
            f"Không thể ép kiểu '{raw_value}' thành DATETIME (kỳ vọng ISO 8601)"
        )

    if data_type == "JSON":
        if isinstance(raw_value, (dict, list)):
            return raw_value
        try:
            return json.loads(str(raw_value))
        except json.JSONDecodeError as exc:
            raise ValueError(f"Không thể ép kiểu '{raw_value}' thành JSON: {exc}") from exc

    raise ValueError(f"data_type không được hỗ trợ: {data_type}")


def map_and_cast_row(
    raw_row: Dict[str, Any],
    schema_fields: List[Dict[str, Any]],
    field_mapping: Dict[str, str],
) -> Tuple[Dict[str, Any], List[Tuple[str, str]]]:
    """Ánh xạ + ép kiểu 1 dòng thô -> (mapped_fields, errors).

    `errors` là danh sách `(field_name, message)` — có thể rỗng. Trường
    đích không có cột nguồn tương ứng được để `None` (không tính là lỗi;
    việc bắt buộc NOT NULL thuộc phạm vi kiểm tra chất lượng dữ liệu khác,
    UC-038/039).
    """
    data_type_by_field = {f["name"]: f["data_type"] for f in schema_fields}
    reverse_mapping: Dict[str, str] = {}
    for source_col, target_field in field_mapping.items():
        reverse_mapping[target_field] = source_col

    mapped: Dict[str, Any] = {}
    errors: List[Tuple[str, str]] = []
    for target_field, data_type in data_type_by_field.items():
        source_col = reverse_mapping.get(target_field)
        raw_value = raw_row.get(source_col) if source_col is not None else None
        try:
            mapped[target_field] = cast_value(raw_value, data_type)
        except ValueError as exc:
            mapped[target_field] = None
            errors.append((target_field, str(exc)))
    return mapped, errors