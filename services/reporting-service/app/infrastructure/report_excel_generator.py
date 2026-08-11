"""Infrastructure — UC-050 bước 3: "Kết xuất Excel -> Hệ thống trả file".

Dùng `openpyxl` (đã có sẵn trong hệ thống, dùng bởi `data-quality-service`
để đọc Excel ở UC-029 — lần đầu dùng để GHI Excel).
"""
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.entities import ReportFilterConfig, ReportTemplate

_PERIOD_LABELS = {"THANG": "Tháng", "QUY": "Quý", "NAM": "Năm"}


def _format_filters(filters: ReportFilterConfig) -> str:
    parts = [f"Năm {filters.year}"]
    period_label = _PERIOD_LABELS.get(filters.period_type, filters.period_type)
    if filters.period_type == "NAM":
        parts.append(f"Kỳ: {period_label}")
    else:
        parts.append(f"Kỳ: {period_label} {filters.period_value}")
    if filters.org_unit_code:
        parts.append(f"Đơn vị: {filters.org_unit_code}")
    if filters.sector:
        parts.append(f"Lĩnh vực: {filters.sector}")
    return " | ".join(parts)


class OpenpyxlReportExcelGenerator:
    """Implement cổng sinh Excel (.xlsx) cho 1 báo cáo đã sinh theo mẫu + bộ lọc (UC-050)."""

    def generate(
        self,
        template: ReportTemplate,
        filters: ReportFilterConfig,
        rows: List[Dict[str, Any]],
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = (template.code or "BaoCao")[:31]

        columns = template.columns
        n_cols = max(len(columns), 1)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=template.name)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="left")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.cell(row=2, column=1, value=_format_filters(filters)).font = Font(italic=True)

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
        ws.cell(row=3, column=1, value=f"Tổng số dòng: {len(rows)}")

        header_row = 5
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        for col_index, column in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_index, value=column.get("label", column.get("field")))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_offset, row in enumerate(rows, start=1):
            for col_index, column in enumerate(columns, start=1):
                ws.cell(
                    row=header_row + row_offset,
                    column=col_index,
                    value=row.get(column.get("field")),
                )

        for col_index, column in enumerate(columns, start=1):
            label = str(column.get("label", column.get("field")) or "")
            ws.column_dimensions[get_column_letter(col_index)].width = max(14, len(label) + 4)

        if columns:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}"
                if rows
                else f"A{header_row}:{get_column_letter(len(columns))}{header_row}"
            )
        ws.freeze_panes = f"A{header_row + 1}"

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()