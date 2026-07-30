"""Integration test UC-022: Tiếp nhận file thủ công TABMIS (upload), qua
HTTP API (SQLite in-memory)."""
import io
import os

os.environ["DATABASE_URL"] = "sqlite:///:memory:"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from app.main import app  # noqa: E402

client = TestClient(app)

_SCHEMA_COLUMNS = ["ma_don_vi", "ten_don_vi", "so_tien"]


def _create_data_source(code="UC22-SRC-01", source_system="TABMIS"):
    resp = client.post(
        "/data-sources",
        json={
            "code": code,
            "name": "Nguồn dữ liệu test UC-22",
            "source_system": source_system,
            "provider": "Bộ Tài chính",
            "owner": "Cục CNTT",
        },
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


def _schema_fields():
    return [
        {"name": col, "data_type": "STRING", "nullable": False, "description": ""}
        for col in _SCHEMA_COLUMNS
    ]


def _create_dataset(data_source_id, code="UC22-DS-01"):
    resp = client.post(
        "/datasets",
        json={
            "data_source_id": data_source_id,
            "code": code,
            "name": "Tập dữ liệu test UC-22",
            "schema_fields": _schema_fields(),
        },
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


def _new_tabmis_dataset(prefix):
    data_source_id = _create_data_source(f"{prefix}-SRC")
    return _create_dataset(data_source_id, f"{prefix}-DS")


def _build_xlsx(columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _upload(dataset_id, content, file_name="tabmis-data.xlsx", uploaded_by="canbo01"):
    return client.post(
        "/tabmis-intake/upload",
        data={"dataset_id": dataset_id, "uploaded_by": uploaded_by},
        files={
            "file": (
                file_name,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


# ---------- Bước 1: Tải biểu mẫu Excel ----------


def test_download_template_returns_columns_from_dataset_schema():
    dataset_id = _new_tabmis_dataset("UC22A")
    resp = client.get("/tabmis-intake/template", params={"dataset_id": dataset_id})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(resp.content))
    header_row = next(wb.active.iter_rows(values_only=True))
    assert list(header_row) == _SCHEMA_COLUMNS


def test_download_template_404_when_dataset_not_found():
    resp = client.get("/tabmis-intake/template", params={"dataset_id": 999999})
    assert resp.status_code == 404
    assert resp.json()["detail"]["code"] == "DATASET_NOT_FOUND"


def test_download_template_422_when_dataset_not_tabmis_source():
    data_source_id = _create_data_source("UC22B-SRC", source_system="MISA")
    dataset_id = _create_dataset(data_source_id, "UC22B-DS")
    resp = client.get("/tabmis-intake/template", params={"dataset_id": dataset_id})
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "DATASET_SOURCE_SYSTEM_MISMATCH"


# ---------- Bước 2-4: Tải tệp lên ----------


def test_upload_valid_file_creates_received_session_and_success_run():
    dataset_id = _new_tabmis_dataset("UC22C")
    content = _build_xlsx(
        _SCHEMA_COLUMNS,
        [["DV01", "Đơn vị 1", "1000000"], ["DV02", "Đơn vị 2", "2000000"]],
    )
    resp = _upload(dataset_id, content)
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["status"] == "RECEIVED"
    assert body["dataset_id"] == dataset_id
    assert body["control_totals"]["records_read"] == 2
    assert body["control_totals"]["missing_columns"] == []
    assert body["error_message"] == ""
    assert body["ingestion_run_id"] is not None
    assert body["raw_object_key"].startswith(f"tabmis-intake/{dataset_id}/")

    run_resp = client.get(f"/ingestion-runs/{body['ingestion_run_id']}")
    assert run_resp.status_code == 200
    run = run_resp.json()
    assert run["status"] == "SUCCESS"
    assert run["records_read"] == 2
    assert run["records_loaded"] == 2
    assert run["trigger"] == "MANUAL"
    assert run["dataset_id"] == dataset_id


def test_upload_missing_column_creates_template_invalid_session_and_failed_run():
    dataset_id = _new_tabmis_dataset("UC22D")
    content = _build_xlsx(
        ["ma_don_vi", "ten_don_vi"],  # thiếu cột so_tien
        [["DV01", "Đơn vị 1"]],
    )
    resp = _upload(dataset_id, content)
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["status"] == "TEMPLATE_INVALID"
    assert "so_tien" in body["control_totals"]["missing_columns"]
    assert body["error_message"] != ""

    run = client.get(f"/ingestion-runs/{body['ingestion_run_id']}").json()
    assert run["status"] == "FAILED"
    assert run["error_message"] != ""


def test_upload_404_when_dataset_not_found():
    content = _build_xlsx(_SCHEMA_COLUMNS, [])
    resp = _upload(999999, content)
    assert resp.status_code == 404
    assert resp.json()["detail"]["code"] == "DATASET_NOT_FOUND"


def test_upload_422_when_dataset_not_tabmis_source():
    data_source_id = _create_data_source("UC22E-SRC", source_system="QLVBDH")
    dataset_id = _create_dataset(data_source_id, "UC22E-DS")
    content = _build_xlsx(_SCHEMA_COLUMNS, [])
    resp = _upload(dataset_id, content)
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "DATASET_SOURCE_SYSTEM_MISMATCH"


def test_upload_422_when_file_extension_invalid():
    dataset_id = _new_tabmis_dataset("UC22F")
    resp = _upload(dataset_id, b"not-an-excel-file", file_name="data.csv")
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "INVALID_TABMIS_INTAKE_UPLOAD"


def test_upload_422_when_file_empty():
    dataset_id = _new_tabmis_dataset("UC22G")
    resp = _upload(dataset_id, b"", file_name="empty.xlsx")
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "INVALID_TABMIS_INTAKE_UPLOAD"


# ---------- Xem lại phiên tiếp nhận ----------


def test_list_sessions_filter_by_dataset_and_status():
    dataset_id = _new_tabmis_dataset("UC22H")
    valid_content = _build_xlsx(_SCHEMA_COLUMNS, [["DV01", "Đơn vị 1", "100"]])
    invalid_content = _build_xlsx(["ma_don_vi"], [["DV01"]])
    _upload(dataset_id, valid_content, file_name="valid.xlsx")
    _upload(dataset_id, invalid_content, file_name="invalid.xlsx")

    all_sessions = client.get("/tabmis-intake", params={"dataset_id": dataset_id}).json()
    assert len(all_sessions) == 2

    received_only = client.get(
        "/tabmis-intake", params={"dataset_id": dataset_id, "status": "RECEIVED"}
    ).json()
    assert len(received_only) == 1
    assert received_only[0]["status"] == "RECEIVED"

    invalid_only = client.get(
        "/tabmis-intake", params={"dataset_id": dataset_id, "status": "TEMPLATE_INVALID"}
    ).json()
    assert len(invalid_only) == 1
    assert invalid_only[0]["status"] == "TEMPLATE_INVALID"


def test_get_session_detail_and_404_when_not_found():
    dataset_id = _new_tabmis_dataset("UC22I")
    content = _build_xlsx(_SCHEMA_COLUMNS, [["DV01", "Đơn vị 1", "100"]])
    created = _upload(dataset_id, content).json()

    resp = client.get(f"/tabmis-intake/{created['id']}")
    assert resp.status_code == 200
    assert resp.json()["id"] == created["id"]

    resp_404 = client.get("/tabmis-intake/999999")
    assert resp_404.status_code == 404
    assert resp_404.json()["detail"]["code"] == "TABMIS_INTAKE_SESSION_NOT_FOUND"