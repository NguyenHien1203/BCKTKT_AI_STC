"""Triển khai ExcelTemplateValidator (UC-022) dùng thư viện `openpyxl`.

- `build_template(columns)`: sinh tệp .xlsx chỉ có 1 dòng tiêu đề đúng thứ
  tự `columns` (tên các trường theo lược đồ dataset TABMIS) — dùng cho bước
  "Tải biểu mẫu Excel" (hệ thống trả về tệp biểu mẫu chuẩn).
- `validate(content, expected_columns)`: đọc dòng tiêu đề của tệp tải lên,
  đối chiếu với `expected_columns`; đồng thời đếm số dòng dữ liệu (không
  tính dòng tiêu đề, bỏ qua dòng trống hoàn toàn) để làm 1 phần tổng kiểm
  soát (control totals).
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

from app.domain.entities import TemplateValidationResult
from app.domain.repositories import ExcelTemplateValidator


class OpenpyxlExcelTemplateValidator(ExcelTemplateValidator):
    SHEET_NAME = "TABMIS"

    def build_template(self, columns: List[str]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.SHEET_NAME
        sheet.append(list(columns))
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def validate(self, content: bytes, expected_columns: List[str]) -> TemplateValidationResult:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # tệp hỏng/không phải Excel hợp lệ
            return TemplateValidationResult(
                valid=False,
                message=f"Không đọc được tệp Excel: {exc}",
                found_columns=[],
                missing_columns=list(expected_columns),
                row_count=0,
            )

        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            header_row = ()

        found_columns = [str(cell).strip() for cell in header_row if cell is not None]
        missing_columns = [col for col in expected_columns if col not in found_columns]
        row_count = sum(
            1 for row in rows_iter if any(cell is not None and str(cell).strip() for cell in row)
        )

        valid = len(missing_columns) == 0
        message = (
            "Tệp đúng biểu mẫu chuẩn"
            if valid
            else f"Tệp thiếu cột bắt buộc: {', '.join(missing_columns)}"
        )
        return TemplateValidationResult(
            valid=valid,
            message=message,
            found_columns=found_columns,
            missing_columns=missing_columns,
            row_count=row_count,
        )

    # ---------- UC-023 bước 2: kiểm tra lỗi từng dòng dữ liệu ----------

    _NUMERIC_TYPES = {"INTEGER", "BIGINT", "DECIMAL"}
    _DATE_TYPES = {"DATE", "DATETIME"}
    _BOOLEAN_TRUE = {"true", "1", "co", "có"}
    _BOOLEAN_FALSE = {"false", "0", "khong", "không"}

    @classmethod
    def _cell_error(cls, data_type: str, value: Any) -> str:
        """Trả về thông báo lỗi nếu `value` không khớp `data_type`, hoặc
        chuỗi rỗng nếu hợp lệ."""
        if data_type in cls._NUMERIC_TYPES:
            if isinstance(value, bool):
                return f"phải là số ({data_type}), không phải kiểu luận lý"
            if isinstance(value, (int, float)):
                return ""
            try:
                float(str(value).strip().replace(",", ""))
                return ""
            except (TypeError, ValueError):
                return f"giá trị '{value}' không phải kiểu số ({data_type})"
        if data_type == "BOOLEAN":
            if isinstance(value, bool):
                return ""
            text = str(value).strip().lower()
            if text in cls._BOOLEAN_TRUE or text in cls._BOOLEAN_FALSE:
                return ""
            return f"giá trị '{value}' không phải kiểu luận lý (BOOLEAN)"
        if data_type in cls._DATE_TYPES:
            if isinstance(value, (date, datetime)):
                return ""
            text = str(value).strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
                try:
                    datetime.strptime(text, fmt)
                    return ""
                except ValueError:
                    continue
            return f"giá trị '{value}' không đúng định dạng ngày/giờ ({data_type})"
        # STRING/JSON: không kiểm tra định dạng, chỉ kiểm tra NOT NULL ở nơi gọi
        return ""

    def validate_rows(
        self,
        content: bytes,
        schema_fields: List[Dict[str, Any]],
        critical_field_names: List[str],
    ) -> List[Dict[str, Any]]:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception:
            # Tệp không đọc được: đã được `validate()` báo lỗi biểu mẫu ở
            # bước trước, không cần lặp lại lỗi ở đây.
            return []

        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        header = [str(cell).strip() if cell is not None else "" for cell in header_row]
        critical_set = set(critical_field_names)
        field_data_type = {f["name"]: f.get("data_type", "STRING") for f in schema_fields}

        errors: List[Dict[str, Any]] = []
        row_number = 0
        for row in rows_iter:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue  # bỏ qua dòng trống hoàn toàn
            row_number += 1
            row_values = dict(zip(header, row))
            for field_name, data_type in field_data_type.items():
                value = row_values.get(field_name)
                is_empty = value is None or (isinstance(value, str) and not value.strip())
                if is_empty:
                    if field_name in critical_set:
                        errors.append(
                            {
                                "row_number": row_number,
                                "field_name": field_name,
                                "message": f"Trường bắt buộc '{field_name}' bị bỏ trống",
                            }
                        )
                    continue
                type_error = self._cell_error(data_type, value)
                if type_error:
                    errors.append(
                        {
                            "row_number": row_number,
                            "field_name": field_name,
                            "message": f"Trường '{field_name}' {type_error}",
                        }
                    )
        return errors